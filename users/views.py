import openpyxl
from django.views.decorators.http import require_POST
from rest_framework import generics, permissions
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.db.models import Count, Sum

from . import serializers

from events.models import EventRegistration, BonusTaskCompletion, Event
from .models import (
    User, WalletTransaction, BonusTask, BonusCompletion,
    Prize, PrizeRedemption, QuizQuestion, QuizAnswer, BonusTransaction,
)
from .serializers import RegisterSerializer, UserSerializer
from .forms import RegisterForm, LoginForm, BonusTaskForm, QuizQuestionFormSet, CommentForm
from .utils import send_verification_email

from decimal import Decimal


class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = RegisterSerializer
    permission_classes = [permissions.AllowAny]


class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        if not user.is_verified:
            raise serializers.ValidationError("Ваш email не подтверждён. Проверьте почту.")
        token = super().get_token(user)
        token['role'] = user.role
        return token


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


class UserListView(generics.ListCreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]


class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]


User = get_user_model()


def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False
            user.save()

            send_verification_email(user)

            task = BonusTask.objects.filter(
                type=BonusTask.SYSTEM,
                is_active=True,
                code__isnull=True,
                name__icontains="регистрация"
            ).first()

            if task:
                user.bonus_balance += task.reward
                user.save()

                from .models import BonusTransaction
                BonusTransaction.objects.create(
                    user=user,
                    amount=task.reward,
                    type=BonusTransaction.INCOME,
                    description="Бонус за регистрацию"
                )
                BonusCompletion.objects.create(user=user, task=task)

            return render(request, "users/verification_sent.html")
    else:
        form = RegisterForm()

    return render(request, "users/register.html", {"form": form})


def verify_email(request, token):
    user = get_object_or_404(User, verification_token=token)
    user.is_verified = True
    user.is_active = True
    user.verification_token = None
    user.save()

    messages.success(request, "Ваш email подтверждён. Теперь можно войти.")
    return redirect("login")


def user_login(request):
    if request.method == "POST":
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            if not user.is_verified:
                messages.error(request, "Email не подтверждён.")
                return redirect("login")
            login(request, user)
            return redirect("event-list")
    else:
        form = LoginForm()

    return render(request, "users/login.html", {"form": form})


@login_required
def wallet_view(request):
    transactions = request.user.transactions.all()

    if request.method == "POST":
        try:
            amount = Decimal(request.POST.get("amount"))
            if amount <= 0:
                messages.error(request, "Сумма должна быть положительной.")
            else:
                request.user.wallet_balance += amount
                request.user.save()

                WalletTransaction.objects.create(
                    user=request.user,
                    amount=amount,
                    type=WalletTransaction.INCOME,
                    description="Пополнение вручную"
                )
                messages.success(request, f"Баланс пополнен на {amount}₽.")
                return redirect("wallet")
        except (ValueError, TypeError):
            messages.error(request, "Введите корректную сумму.")

    return render(request, "users/wallet.html", {
        "balance": request.user.wallet_balance,
        "transactions": transactions
    })


@login_required
def wallet_history(request):
    transactions = WalletTransaction.objects.filter(user=request.user)
    return render(request, 'users/wallet_history.html', {'transactions': transactions})


@login_required
def admin_dashboard(request):
    """Админ-панель с обзором пользователей, транзакций и бонусных задач."""
    if not request.user.is_admin():
        messages.error(request, "Доступ запрещён.")
        return redirect("event-list")

    users_count = User.objects.count()
    transactions_count = WalletTransaction.objects.count()
    bonus_tasks_count = BonusTask.objects.count()

    return render(request, "admin/dashboard.html", {
        "users_count": users_count,
        "transactions_count": transactions_count,
        "bonus_tasks_count": bonus_tasks_count,
    })


@login_required
def manage_bonus_tasks(request):
    """Просмотр и управление бонусными задачами (только для админов)."""
    if not request.user.is_admin():
        messages.error(request, "Доступ запрещён.")
        return redirect("admin-dashboard")

    bonus_tasks = BonusTask.objects.all()

    return render(request, "admin/manage_bonus_tasks.html", {
        "bonus_tasks": bonus_tasks
    })


@login_required
def bonus_task_create(request):
    if not request.user.is_admin() and not request.user.is_organizer():
        messages.error(request, "Доступ запрещён.")
        return redirect("event-list")

    event_id = request.GET.get("event")
    event = None
    if event_id:
        event = get_object_or_404(Event, id=event_id)
        if not request.user.is_superuser and event.organizer != request.user:
            messages.error(request, "У вас нет прав на добавление задания к этому мероприятию.")
            return redirect("event-detail", pk=event.id)

    if request.method == "POST":
        form = BonusTaskForm(request.POST)
        question_formset = QuizQuestionFormSet(request.POST)
        if form.is_valid() and (form.cleaned_data["type"] != "quiz" or question_formset.is_valid()):
            task = form.save(commit=False)
            if event:
                task.event = event
            task.save()
            if task.type == "quiz":
                for q_form in question_formset:
                    if q_form.cleaned_data:
                        question = q_form.save(commit=False)
                        question.task = task
                        question.save()
            messages.success(request, "Бонусное задание создано!")
            if event:
                return redirect("event-bonus-tasks", event_id=event.id)
            return redirect("available-bonus-tasks")
    else:
        form = BonusTaskForm()
        question_formset = QuizQuestionFormSet(queryset=QuizQuestion.objects.none())

    return render(request, "users/bonus_task_form_user.html", {
        "form": form,
        "is_edit": False,
        "event": event,
        "question_formset": question_formset
    })


@login_required
def bonus_task_edit(request, task_id):
    """Редактирование бонусного задания."""
    task = get_object_or_404(BonusTask, id=task_id)

    if not request.user.is_admin():
        messages.error(request, "Доступ запрещён.")
        return redirect("available-bonus-tasks")

    if request.method == "POST":
        form = BonusTaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            messages.success(request, "Задание обновлено!")
            return redirect("available-bonus-tasks")
    else:
        form = BonusTaskForm(instance=task)

    return render(request, "admin/bonus_task_form.html", {"form": form, "is_edit": True, "task": task})


@login_required
def bonus_task_delete(request, task_id):
    """Удаление бонусного задания."""
    task = get_object_or_404(BonusTask, id=task_id)

    if not request.user.is_admin():
        messages.error(request, "Доступ запрещён.")
        return redirect("available-bonus-tasks")

    task.delete()
    messages.success(request, "Задание удалено.")
    return redirect("available-bonus-tasks")


@login_required
def user_profile(request):
    """Профиль пользователя: регистрация на события, выполненные задания, баланс."""
    registered_events = EventRegistration.objects.filter(user=request.user).select_related('event')
    bonus_tasks = BonusTaskCompletion.objects.filter(user=request.user).select_related('task')

    return render(request, 'users/profile.html', {
        'registered_events': registered_events,
        'bonus_tasks': bonus_tasks,
        'balance': request.user.wallet_balance
    })


def prize_catalog(request):
    """Публичный список призов"""
    prizes = Prize.objects.filter(is_active=True)
    return render(request, 'users/prize_catalog.html', {'prizes': prizes})


@login_required
def redeem_prize(request, prize_id):
    prize = get_object_or_404(Prize, id=prize_id, is_active=True)

    if request.user.bonus_balance < prize.cost:
        messages.error(request, "Недостаточно бонусов для получения приза.")
        return redirect('prize-catalog')

    request.user.bonus_balance -= prize.cost
    request.user.save()

    PrizeRedemption.objects.create(user=request.user, prize=prize)

    from .models import BonusTransaction
    BonusTransaction.objects.create(
        user=request.user,
        type=BonusTransaction.EXPENSE,
        amount=prize.cost,
        description=f"Получен приз: {prize.name}"
    )

    messages.success(request, f"Вы получили приз: {prize.name}!")
    return redirect('prize-catalog')

    request.user.wallet_balance -= prize.cost
    request.user.save()

    PrizeRedemption.objects.create(user=request.user, prize=prize)

    WalletTransaction.objects.create(
        user=request.user,
        type=WalletTransaction.EXPENSE,
        amount=prize.cost,
        description=f"Получен приз: {prize.name}"
    )

    messages.success(request, f"Вы получили приз: {prize.name}!")
    return redirect('prize-catalog')


@login_required
def my_prizes(request):
    """Показывает список призов, которые получил пользователь"""
    redemptions = PrizeRedemption.objects.filter(user=request.user).select_related('prize').order_by('-redeemed_at')
    return render(request, 'users/my_prizes.html', {'redemptions': redemptions})


@login_required
@require_POST
def submit_bonus_code(request):
    code = request.POST.get("code", "").strip()

    if not code:
        messages.error(request, "Пожалуйста, введите код.")
        return redirect("user-profile")

    try:
        task = BonusTask.objects.get(code__iexact=code, is_active=True)
    except BonusTask.DoesNotExist:
        messages.error(request, "Код недействителен или уже использован.")
        return redirect("user-profile")

    if BonusTaskCompletion.objects.filter(user=request.user, task=task).exists():
        messages.info(request, "Вы уже использовали этот код.")
        return redirect("user-profile")

    request.user.bonus_balance += task.reward
    request.user.save()

    BonusTaskCompletion.objects.create(user=request.user, task=task)

    from .models import BonusTransaction
    BonusTransaction.objects.create(
        user=request.user,
        amount=task.reward,
        type=BonusTransaction.INCOME,
        description=f"Бонус за код: {task.name}"
    )

    messages.success(request, f"Поздравляем! Вы получили {task.reward} бонусов за задание «{task.name}».")
    return redirect("user-profile")  # или другую страницу

    try:
        task = BonusTask.objects.get(code__iexact=code, is_active=True)
    except BonusTask.DoesNotExist:
        messages.error(request, "Код недействителен или уже использован.")
        return redirect("user-profile")

    if BonusTaskCompletion.objects.filter(user=request.user, task=task).exists():
        messages.info(request, "Вы уже использовали этот код.")
        return redirect("user-profile")

    request.user.bonus_balance += task.reward
    request.user.save()

    BonusTaskCompletion.objects.create(user=request.user, task=task)
    messages.success(request, f"Поздравляем! Вы получили {task.reward}₽ за задание «{task.name}».")

    return redirect("user-profile")


@login_required
def available_bonus_tasks(request):
    """Список доступных заданий (глобальных и для мероприятий, на которые пользователь записан)"""
    completed_ids = BonusTaskCompletion.objects.filter(user=request.user).values_list('task_id', flat=True)

    registered_event_ids = Event.objects.filter(eventregistration__user=request.user).values_list('id', flat=True)

    global_tasks = BonusTask.objects.filter(is_active=True, event__isnull=True)

    event_tasks = BonusTask.objects.filter(is_active=True, event_id__in=registered_event_ids)

    all_tasks = (global_tasks | event_tasks).distinct().order_by('-reward')

    return render(request, 'users/available_tasks.html', {
        'tasks': all_tasks,
        'completed_ids': set(completed_ids)
    })


@login_required
def bonus_task_create(request):
    """Создание бонусного задания (с возможной привязкой к мероприятию)."""
    if not request.user.is_admin() and not request.user.is_organizer():
        messages.error(request, "Доступ запрещён.")
        return redirect("event-list")

    event_id = request.GET.get("event")
    event = None

    if event_id:
        from events.models import Event
        event = get_object_or_404(Event, id=event_id)

        # Только организатор этого мероприятия или админ
        if not request.user.is_superuser and event.organizer != request.user:
            messages.error(request, "У вас нет прав на добавление задания к этому мероприятию.")
            return redirect("event-detail", pk=event.id)

    if request.method == "POST":
        form = BonusTaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.event = event  # ← Привязываем мероприятие, если было передано
            task.save()
            messages.success(request, "Бонусное задание создано!")
            if event:
                return redirect("event-bonus-tasks", event_id=event.id)
            return redirect("available-bonus-tasks")
    else:
        form = BonusTaskForm()

    return render(request, "users/bonus_task_form_user.html", {
        "form": form,
        "is_edit": False,
        "event": event
    })


@login_required
def bonus_task_public(request, code):
    try:
        task = BonusTask.objects.get(code__iexact=code, is_active=True)
    except BonusTask.DoesNotExist:
        messages.error(request, "Задание не найдено или недействительно.")
        return redirect("event-list")

    already_done = BonusTaskCompletion.objects.filter(user=request.user, task=task).exists()

    if request.method == "POST" and not already_done:
        request.user.bonus_balance += task.reward
        request.user.save()

        BonusTaskCompletion.objects.create(user=request.user, task=task)
        messages.success(request, f"Вы получили {task.reward}₽ за выполнение задания!")
        return redirect("bonus-task-public", code=code)

    return render(request, "users/bonus_task_public.html", {
        "task": task,
        "already_done": already_done
    })


def public_bonus_tasks(request):
    """Показывает активные задания с QR-кодами (публично)."""
    tasks = BonusTask.objects.filter(is_active=True, code__isnull=False)
    return render(request, 'users/available_tasks.html', {'tasks': tasks})


@login_required
def event_user_tasks(request, event_id):
    """Показывает задания для мероприятия и статус их выполнения текущим пользователем."""
    event = get_object_or_404(Event, id=event_id)

    if not event.enable_tasks:
        messages.warning(request, "На этом мероприятии система заданий отключена.")
        return redirect('event-detail', pk=event_id)

    tasks = BonusTask.objects.filter(event=event, is_active=True)
    completed_ids = BonusTaskCompletion.objects.filter(user=request.user, task__in=tasks).values_list('task_id',
                                                                                                      flat=True)

    return render(request, 'users/event_tasks.html', {
        'event': event,
        'tasks': tasks,
        'completed_ids': set(completed_ids),
    })


@login_required
def task_completions_view(request, task_id):
    task = get_object_or_404(BonusTask, id=task_id)
    User = get_user_model()

    # Защита: только организатор задания или админ
    if task.event:
        is_owner = task.event.organizer == request.user
    else:
        is_owner = request.user.is_admin()

    if not (is_owner or request.user.is_superuser):
        messages.error(request, "У вас нет доступа к просмотру этого задания.")
        return redirect("event-list")

    completions = BonusTaskCompletion.objects.filter(task=task).select_related("user").order_by("-completed_at")

    return render(request, "admin/task_completions.html", {
        "task": task,
        "completions": completions,
    })


@login_required
def export_task_completions_xlsx(request, task_id):
    task = get_object_or_404(BonusTask, id=task_id)

    # Доступ только организатору или админу
    if task.event:
        is_owner = task.event.organizer == request.user
    else:
        is_owner = request.user.is_admin()

    if not (is_owner or request.user.is_superuser):
        messages.error(request, "У вас нет доступа к этому заданию.")
        return redirect("event-list")

    completions = BonusTaskCompletion.objects.filter(task=task).select_related("user").order_by("completed_at")

    # Создание Excel-файла
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Выполнения"

    headers = ["Имя пользователя", "Email", "Баланс (₽)", "Дата выполнения"]
    ws.append(headers)

    for comp in completions:
        ws.append([
            comp.user.username,
            comp.user.email,
            float(comp.user.balance),
            comp.completed_at.strftime("%d.%m.%Y %H:%M"),
        ])

    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"{task.name.replace(' ', '_')}_выполнения.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


@login_required
def export_event_completions_xlsx(request, event_id):
    from events.models import BonusTask, BonusTaskCompletion

    event = get_object_or_404(Event, id=event_id)

    if request.user != event.organizer and not request.user.is_superuser:
        messages.error(request, "У вас нет доступа к этому мероприятию.")
        return redirect("event-detail", pk=event.id)

    tasks = BonusTask.objects.filter(event=event)
    completions = BonusTaskCompletion.objects.filter(task__in=tasks).select_related("user", "task").order_by(
        "task__name", "completed_at")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Все выполнения"

    headers = ["Задание", "Пользователь", "Email", "Награда (₽)", "Дата выполнения"]
    ws.append(headers)

    for c in completions:
        ws.append([
            c.task.name,
            c.user.username,
            c.user.email,
            float(c.task.reward),
            c.completed_at.strftime("%d.%m.%Y %H:%M"),
        ])

    for col_num, column_title in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"{event.title.replace(' ', '_')}_выполнения.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


def leaderboard_view(request):
    from django.contrib.auth import get_user_model
    User = get_user_model()

    users = User.objects.annotate(total_tasks=Count('bonustaskcompletion')).order_by('-bonus_balance', '-total_tasks')[:20]

    return render(request, 'users/leaderboard.html', {
        'leaders': users,
    })


@login_required
def pass_quiz(request, task_id):
    task = get_object_or_404(BonusTask, id=task_id, type='quiz', is_active=True)
    questions = task.questions.all()

    already_done = BonusTaskCompletion.objects.filter(user=request.user, task=task).exists()

    if request.method == "POST" and not already_done:
        correct = 0
        for q in questions:
            user_answer = request.POST.get(f"q_{q.id}", "").strip()
            QuizAnswer.objects.create(user=request.user, question=q, answer=user_answer)
            if user_answer.lower() == q.correct_answer.lower():
                correct += 1

        if correct == len(questions):
            request.user.bonus_balance += task.reward
            request.user.save()
            BonusTaskCompletion.objects.create(user=request.user, task=task)
            WalletTransaction.objects.create(
                user=request.user,
                amount=task.reward,
                type=BonusTransaction.INCOME,
                description=f"Викторина: {task.name}"
            )
            messages.success(request, f"Поздравляем! Вы прошли викторину и получили {task.reward}₽.")
        else:
            messages.warning(request, f"Викторина не пройдена. Верных ответов: {correct} из {len(questions)}.")

        return redirect("user-profile")

    return render(request, "users/pass_quiz.html", {
        "task": task,
        "questions": questions,
        "already_done": already_done
    })


@login_required
def ticket_list(request):
    if not request.user.is_admin():
        messages.error(request, "Доступ запрещён.")
        return redirect("user-profile")
    from .models import Ticket
    ticket_type = request.GET.get("type")
    tickets = Ticket.objects.all().order_by("-created_at")
    status = request.GET.get("status")
    if status in ["open", "closed"]:
        tickets = tickets.filter(status=status)
    if ticket_type in ["question", "report", "promotion"]:
        tickets = tickets.filter(type=ticket_type)
    return render(request, "admin/ticket_list.html", {"tickets": tickets})


@login_required
def ticket_detail(request, ticket_id):
    if not request.user.is_admin():
        messages.error(request, "Доступ запрещён.")
        return redirect("user-profile")
    from .models import Ticket, OrganizerApplication, User
    ticket = get_object_or_404(Ticket, id=ticket_id)

    if request.method == "POST":
        action = request.POST.get("action")
        response = request.POST.get("response", "").strip()

        if response:
            ticket.admin_response = response

        if action == "close":
            ticket.status = "closed"
            ticket.save()
            messages.success(request, "Тикет закрыт.")
        elif action == "accept" and ticket.type == "promotion":
            ticket.status = "closed"
            ticket.save()
            ticket.created_by.role = User.ORGANIZER
            ticket.created_by.save()
            messages.success(request, "Пользователь назначен организатором.")
        elif action == "reject" and ticket.type == "promotion":
            ticket.status = "closed"
            ticket.save()
            messages.info(request, "Заявка отклонена.")

        if action:
            ticket.save()
            return redirect("ticket-detail", ticket_id=ticket.id)

    template_map = {
        "promotion": "admin/tickets/detail_promotion.html",
        "report": "admin/tickets/detail_report.html",
        "question": "admin/tickets/detail_question.html",
    }
    reason_text = None
    if ticket.type == "report" and ticket.reported_comment:
        if "Причина:" in ticket.comment:
            reason_text = ticket.comment.split("Причина:", 1)[-1].strip()
    return render(request, template_map.get(ticket.type, "admin/ticket_detail.html"), {
        "ticket": ticket,
        "reason_text": reason_text
    })


@login_required
def apply_organizer(request):
    if request.user.is_organizer():
        messages.info(request, "Вы уже являетесь организатором.")
        return redirect("user-profile")

    from .models import OrganizerApplication, Ticket
    from .forms import OrganizerApplicationForm

    from .models import Ticket
    if Ticket.objects.filter(created_by=request.user, type="promotion", status="open").exists():
        messages.info(request, "Заявка уже отправлена и находится на рассмотрении.")
        return redirect("user-profile")

    if request.method == "POST":
        form = OrganizerApplicationForm(request.POST)
        if form.is_valid():
            application = form.save(commit=False)
            application.user = request.user
            application.save()

            Ticket.objects.create(
                created_by=request.user,
                type='promotion',
                comment="Заявка на получение статуса организатора"
            )

            messages.success(request, "Заявка отправлена. Ожидайте решения администрации.")
            return redirect("user-profile")
    else:
        form = OrganizerApplicationForm()

    return render(request, "users/apply_organizer.html", {"form": form})


@login_required
def my_tickets(request):
    from .models import Ticket
    tickets = Ticket.objects.filter(created_by=request.user).order_by('-created_at')
    return render(request, "users/my_tickets.html", {"tickets": tickets})


@login_required
def ticket_conversation(request, ticket_id):
    from .models import Ticket, TicketMessage
    from .forms import TicketMessageForm
    ticket = get_object_or_404(Ticket, id=ticket_id)

    if request.user != ticket.created_by and not request.user.is_admin():
        messages.error(request, "Доступ запрещён.")
        return redirect("user-profile")

    if request.method == "POST" and ticket.status != "closed":
        form = TicketMessageForm(request.POST)
        if form.is_valid():
            msg = form.save(commit=False)
            msg.ticket = ticket
            msg.sender = request.user
            msg.save()

            # Уведомление на почту другой стороне
            from django.core.mail import send_mail
            recipient = ticket.created_by.email if request.user.is_admin() else "admin@example.com"  # заменить на реальный адрес
            send_mail(
                subject=f"Новое сообщение по тикету #{ticket.id}",
                message=msg.text,
                from_email=None,
                recipient_list=[recipient],
                fail_silently=True,
            )

            return redirect("ticket-conversation", ticket_id=ticket.id)
    else:
        form = TicketMessageForm()

    messages_qs = TicketMessage.objects.filter(ticket=ticket).select_related("sender").order_by("created_at")

    return render(request, "users/ticket_conversation.html", {
        "ticket": ticket,
        "messages": messages_qs,
        "form": form
    })


@login_required
def create_ticket(request):
    from .forms import TicketForm
    from .models import Ticket
    if request.method == "POST":
        form = TicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            ticket.type = form.cleaned_data['type']
            ticket.save()
            messages.success(request, "Тикет отправлен администрации.")
            return redirect("my-tickets")
    else:
        form = TicketForm()
    return render(request, "users/create_ticket.html", {"form": form})


@login_required
def report_comment(request, comment_id):
    from .models import Comment, Ticket
    from .forms import ReportCommentForm

    comment = get_object_or_404(Comment, id=comment_id)

    if request.method == "POST":
        form = ReportCommentForm(request.POST)
        if form.is_valid():
            reason = form.cleaned_data["reason"]
            text = f"""Жалоба на комментарий:

"{comment.content}"

Причина: {reason}"""

            Ticket.objects.create(
                type='report',
                created_by=request.user,
                reported_comment=comment,
                comment=text,
                status='open'
            )

            messages.success(request, "Жалоба отправлена администрации.")
            return redirect("event-detail", event_id=comment.event.id)
    else:
        form = ReportCommentForm()

    return render(request, "users/report_comment.html", {
        "form": form,
        "comment": comment
    })


@login_required
def create_ticket_question(request):
    from .forms import TicketForm
    if request.method == "POST":
        form = TicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            ticket.type = 'question'
            ticket.save()
            messages.success(request, "Вопрос отправлен администрации.")
            return redirect("my-tickets")
    else:
        form = TicketForm()
    return render(request, "users/create_ticket_question.html", {"form": form})


@login_required
def create_ticket_report_user(request):
    from .forms import TicketForm
    if request.method == "POST":
        form = TicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            ticket.type = 'report'
            ticket.save()
            messages.success(request, "Жалоба отправлена администрации.")
            return redirect("my-tickets")
    else:
        form = TicketForm()
    return render(request, "users/create_ticket_report_user.html", {"form": form})


from django.views.decorators.http import require_http_methods
from django.db.models import Q
from .models import User
from django.contrib.auth.decorators import user_passes_test


@user_passes_test(lambda u: u.is_admin())
@login_required
def admin_user_list_view(request):
    query = request.GET.get("q", "")
    banned = request.GET.get("banned") == "1"

    users = User.objects.all()

    if query:
        if query.isdigit():
            users = users.filter(Q(id=int(query)) | Q(username__icontains=query))
        else:
            users = users.filter(username__icontains=query)

    if banned:
        users = users.filter(is_active=False)

    role_choices = User.ROLE_CHOICES
    return render(request, "admin/admin_user_list.html", {
        "users": users.order_by("id"),
        "role_choices": role_choices
    })


@require_http_methods(["POST"])
@user_passes_test(lambda u: u.is_admin())
def admin_change_role(request, user_id):
    user = get_object_or_404(User, id=user_id)
    role = request.POST.get("role")
    if role in dict(User.ROLE_CHOICES):
        user.role = role
        user.save()
        messages.success(request, f"Роль обновлена: {user.username} → {role}")
    return redirect("admin-user-list")


@require_http_methods(["POST"])
@user_passes_test(lambda u: u.is_admin())
def admin_toggle_ban(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = not user.is_active
    user.save()
    status = "разбанен" if user.is_active else "забанен"
    messages.info(request, f"Пользователь {user.username} {status}.")
    return redirect("admin-user-list")
